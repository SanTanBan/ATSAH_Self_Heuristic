import winsound
import os
import openpyxl
from ATSAH import ATS_Algorithm_Heuristic
from Gurobi_Algorithm import Gurobi_Algorithm
from PuLP_Algorithm import PuLP_Algorithm
import GeneRead
"""
num_of_instances=7
array_of_instances = [i for i in range(num_of_instances)]

# Call a Workbook() function of openpyxl to create a new blank Workbook object
Combined_Solution_Comparison= openpyxl.Workbook()
Solo_Sheet = Combined_Solution_Comparison.create_sheet("ATSAH Finest")

#Objective_Sheet = Combined_Solution_Comparison.create_sheet("Objectives")
#Time_Sheet = Combined_Solution_Comparison.create_sheet("Solution Time")
#Vehicle_Sheet = Combined_Solution_Comparison.create_sheet("Used Vehicles")
#Total_Iterations = Combined_Solution_Comparison.create_sheet("No. of Iterations")
#Per_Iteration_Span = Combined_Solution_Comparison.create_sheet("100 Iteration Span")
del Combined_Solution_Comparison["Sheet"]

# Providing Column Headings
cell = Solo_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = Solo_Sheet.cell(row = 1, column = 2)
cell.value = "Objectives"
cell = Solo_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 4)
cell.value = "Used Vehicles"
cell = Solo_Sheet.cell(row = 1, column = 5)
cell.value = "No. of Iterations"
cell = Solo_Sheet.cell(row = 1, column = 6)
cell.value = "Per Iteration Span"

for instance in array_of_instances:

    instance_dir_name="Instance "+str(instance)
    os.mkdir(instance_dir_name)
    instance_dir_name=instance_dir_name+"/"

    print("\n #+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#")
    print("~*~*~*~*~*~*~*~*~*~*~*~*~*~*~*~* Starting Instance:",instance," *~*~*~*~*~*~*~*~*~*~*~*~*~*~*~*~")
    print("#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+#+# \n")


    ATSAH_Solution,ATSAH_Time,Decoding_Counts=ATS_Algorithm_Heuristic(directory_details_for_saving=instance_dir_name)
    winsound.Beep(500, 750)
    print(ATSAH_Solution,"\n",ATSAH_Time)

    cell = Solo_Sheet.cell(row = instance+2, column = 2)
    cell.value = ATSAH_Solution[0]

    cell = Solo_Sheet.cell(row =instance+2, column = 3)
    cell.value = ATSAH_Time

    cell = Solo_Sheet.cell(row = instance+2, column = 4)
    vehicle_counter=0
    for i in ATSAH_Solution[3]:
        vehicle_counter+=ATSAH_Solution[3][i]
    cell.value = vehicle_counter

    cell = Solo_Sheet.cell(row = instance+2, column = 5)
    cell.value = Decoding_Counts

    cell = Solo_Sheet.cell(row = instance+2, column = 6)
    if Decoding_Counts==0:
        Decoding_Counts=-1
    cell.value = ATSAH_Time/Decoding_Counts
    
    # sAVING ALL sOLUTIONS oBTAINED
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


for i in range(3):
    winsound.Beep(425, 125)
    winsound.Beep(575, 175)


objec,Vehicle_Type_Max_Utilised_Cap,Vehicles_used_each_Type=PuLP_Algorithm(max_seconds_allowed_for_calculation=9999)
print(objec,Vehicle_Type_Max_Utilised_Cap,Vehicles_used_each_Type)

for i in range(3):
    winsound.Beep(425, 125)
    winsound.Beep(575, 175)


"""
objec,Vehicle_Type_Max_Utilised_Cap,Vehicles_used_each_Type=Gurobi_Algorithm(max_seconds_allowed_for_calculation=6666)
print(objec,Vehicle_Type_Max_Utilised_Cap,Vehicles_used_each_Type)

for i in range(3):
    winsound.Beep(425, 125)
    winsound.Beep(575, 175)


#GeneRead.Generator.Node_Generator(250,latitude_uniform_distribution_upper_bound=50,pickup_quantity_uniform_distribution_upper_bound=100,longitude_uniform_distribution_upper_bound=50,delivery_quantity_uniform_distribution_upper_bound=100)
#GeneRead.Generator.Lp_Norm_Random_Matrix_for_each_Vehicle_Type_Generator(p=2,Vehicle_Types=[2,4])